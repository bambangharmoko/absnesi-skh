import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime, date
import calendar
from typing import Optional
from backend.app.db.database import get_db
from backend.app.db.models import Student, Attendance
from backend.app.core.config import settings

router = APIRouter()

@router.get("/export-excel")
def export_attendance_excel(
    month: Optional[int] = Query(None, ge=1, le=12),
    year: Optional[int] = Query(None, ge=2020, le=2035),
    class_name: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Export comprehensive monthly attendance spreadsheet (.xlsx) for SKH.
    """
    current_date = date.today()
    target_month = month or current_date.month
    target_year = year or current_date.year
    month_name = calendar.month_name[target_month]
    num_days = calendar.monthrange(target_year, target_month)[1]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rekap {month_name} {target_year}"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    title_font = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Arial", size=11, italic=True, color="475569")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    status_fills = {
        "H": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"), # Light Green
        "T": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"), # Light Yellow
        "I": PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"), # Light Blue
        "S": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"), # Light Red
        "A": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"), # Gray
    }

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Header Information
    ws.merge_cells("A1:AJ1")
    ws["A1"] = f"LAPORAN PRESENSI SISWA - {settings.SCHOOL_NAME.upper()}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:AJ2")
    class_label = f"Kelas: {class_name}" if class_name and class_name.lower() != "all" else "Semua Kelas"
    ws["A2"] = f"Periode: {month_name} {target_year} | {class_label} | Dicetak pada: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 26

    # 2. Table Columns
    headers = ["No", "NIS", "Nama Lengkap Siswa", "Kelas", "Kebutuhan Khusus"]
    for day in range(1, num_days + 1):
        headers.append(str(day))
    headers.extend(["Hadir (H)", "Terlambat (T)", "Izin (I)", "Sakit (S)", "Alpha (A)", "% Hadir"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 3. Query Students & Attendances
    student_query = db.query(Student).filter(Student.is_active == True)
    if class_name and class_name.lower() != "all":
        student_query = student_query.filter(Student.class_name == class_name)
    students = student_query.order_by(Student.class_name.asc(), Student.full_name.asc()).all()

    # Query attendances for this month
    start_date_str = f"{target_year:04d}-{target_month:02d}-01"
    end_date_str = f"{target_year:04d}-{target_month:02d}-{num_days:02d}"
    attendances = db.query(Attendance).filter(
        Attendance.date >= start_date_str,
        Attendance.date <= end_date_str
    ).all()

    # Map attendance by student_id and date day
    att_map = {}
    for a in attendances:
        try:
            day_num = int(a.date.split("-")[2])
            att_map[(a.student_id, day_num)] = a.status
        except Exception:
            continue

    # 4. Fill Student Rows
    current_row = 5
    for idx, s in enumerate(students, 1):
        ws.row_dimensions[current_row].height = 20
        row_data = [idx, s.nis, s.full_name, s.class_name, s.category]
        
        count_h = 0
        count_t = 0
        count_i = 0
        count_s = 0
        count_a = 0

        # Day columns
        for day in range(1, num_days + 1):
            st = att_map.get((s.id, day), "")
            symbol = ""
            if st == "HADIR":
                symbol = "H"
                count_h += 1
            elif st == "TERLAMBAT":
                symbol = "T"
                count_t += 1
            elif st == "IZIN":
                symbol = "I"
                count_i += 1
            elif st == "SAKIT":
                symbol = "S"
                count_s += 1
            elif st == "ALPHA":
                symbol = "A"
                count_a += 1
            row_data.append(symbol)

        total_presence = count_h + count_t
        pct = f"{round(total_presence / num_days * 100, 1)}%" if num_days > 0 else "0%"
        row_data.extend([count_h, count_t, count_i, count_s, count_a, pct])

        # Write cells
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            # Alignments
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [3, 4, 5]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Day coloring
            if 6 <= col_idx <= (5 + num_days) and str(val) in status_fills:
                cell.fill = status_fills[str(val)]
                cell.font = bold_font
            elif idx % 2 == 0:
                cell.fill = alt_fill

        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22
    for day_col in range(6, 6 + num_days):
        col_letter = openpyxl.utils.get_column_letter(day_col)
        ws.column_dimensions[col_letter].width = 4.5

    for stat_col in range(6 + num_days, 6 + num_days + 6):
        col_letter = openpyxl.utils.get_column_letter(stat_col)
        ws.column_dimensions[col_letter].width = 12

    # Save to file
    filename = f"rekap_absensi_{target_month}_{target_year}_{int(datetime.utcnow().timestamp())}.xlsx"
    filepath = settings.EXPORTS_DIR / filename
    wb.save(str(filepath))

    return FileResponse(
        str(filepath),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename
    )
